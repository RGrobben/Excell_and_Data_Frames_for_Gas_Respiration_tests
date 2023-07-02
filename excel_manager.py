import dataclasses
import os
import shutil
from datetime import datetime
from typing import List

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

from data_classes import ConstantsSample


class ExcelManager:
    """A class for managing an Excel file."""

    def __init__(self, file_path: str) -> None:
        """
        Initialize the ExcelManager.

        Args:
            file_path (str): The path to the Excel file.
        """
        self.file_path = file_path
        self.workbook_values = None
        self.workbook_formula = None
        self.dict_panda_data_frames = {}
        self.dict_constants_data_frames = {}
        self.dict_constants_data_classes = {}
        self.sheet_names = {}

    def load_workbook(self) -> None:
        """Load the Excel file as a workbook. With data only on True, the workbook is loaded with
        only the values of the cell and not the formulas."""
        self.workbook_values = load_workbook(filename=self.file_path, data_only=True)
        self.workbook_formula = load_workbook(filename=self.file_path)

    def get_sheet_names(self) -> List[str]:
        """
        Get the names of all sheets in the workbook.

        Returns:
            List[str]: A list of sheet names.
        """
        if not self.workbook_values:
            raise Exception("Workbook is not loaded.")
        return self.workbook_values.sheetnames

    def load_sheet_table(
            self,
            sheet_name: str,
            start_row: int,
            start_column: int = None,
            end_row: int = None,
            end_column: int = None,
            values_only: bool = True,
            data_only: bool = False
    ):
        """
            Load a portion of an Excel sheet as a pandas DataFrame.

            Parameters
            ----------
            sheet_name : str
                The name of the sheet to load.
            start_row : int
                The first row of the range to load.
            start_column : int, optional
                The first column of the range to load. If None, starts from the first column.
            end_row : int, optional
                The last row of the range to load. If None, loads until the last row.
            end_column : int, optional
                The last column of the range to load. If None, loads until the last column.
            values_only: bool = True,
                Set True, If you only want to use the workbook with only the values and not also the formulas.
            data_only: bool = False
                Set True, if fou only want the values of the cell and not other additional properties.

            Raises
            ------
            Exception
                If the workbook has not been loaded.

            Returns
            -------
            pd.DataFrame
            """

        if data_only is not None:
            workbook = self.workbook_values
        else:
            workbook = self.workbook_formula

        if not workbook:
            raise Exception("Workbook is not loaded.")

        sheet = workbook[sheet_name]

        data = list(sheet.iter_rows(min_row=start_row,
                                    min_col=start_column,
                                    max_row=end_row,
                                    max_col=end_column,
                                    values_only=values_only))

        header = data[0]
        values = data[1:]
        data_frame = pd.DataFrame(data=values, columns=header)

        return data_frame

    def load_sheet_table_with_input_header(
            self,
            sheet_name: str,
            start_row: int,
            column_names: [str],
            start_column: int = None,
            end_row: int = None,
            end_column: int = None,
            values_only: bool = True,
            data_only: bool = False
    ):
        """
            Load a portion of an Excel sheet as a pandas DataFrame. With your own given list of column names.

            Parameters
            ----------
            sheet_name : str
                The name of the sheet to load.
            start_row : int
                The first row of the range to load.
            column_names: [str]
                A list of the names of the columns
            start_column : int, optional
                The first column of the range to load. If None, starts from the first column.
            end_row : int, optional
                The last row of the range to load. If None, loads until the last row.
            end_column : int, optional
                The last column of the range to load. If None, loads until the last column.
            values_only: bool = True,
                Set True, If you only want to use the workbook with only the values and not also the formulas.
            data_only: bool = False
                Set True, if fou only want the values of the cell and not other additional properties.

            Raises
            ------
            Exception
                If the workbook has not been loaded.

            Returns
            -------
            pd.DataFrame
            """

        if data_only is not None:
            workbook = self.workbook_values
        else:
            workbook = self.workbook_formula

        if not workbook:
            raise Exception("Workbook is not loaded.")

        sheet = workbook[sheet_name]

        data = list(sheet.iter_rows(min_row=start_row,
                                    min_col=start_column,
                                    max_row=end_row,
                                    max_col=end_column,
                                    values_only=values_only))

        values = data[1:]
        data_frame = pd.DataFrame(data=values, columns=column_names)

        return data_frame

    def load_constants_as_data_frame(
            self,
            sheet_name: str,
            start_row: int,
            start_col: int,
            end_row: int = None,
            values_only: bool = True,
            data_only: bool = False
    ) -> pd.DataFrame:
        """
            Load constants from a specified sheet in the workbook and return them as a pandas DataFrame.

            Parameters:
                - sheet_name (str): The name of the sheet containing the constants.
                - start_row (int): The starting row index for reading constants.
                - start_col (int): The starting column index for reading constants.
                - end_row (int, optional): The ending row index for reading constants.
                 Defaults to None, indicating reading until the last row.
                - values_only (bool, optional): Whether to retrieve only the values of cells. Defaults to True.
                - data_only (bool, optional): Whether to load the workbook with only calculated values. Defaults to False.

            Returns:
                pd.DataFrame: A pandas DataFrame containing the loaded constants.

            Raises:
                Exception: If the workbook is not loaded.

            Note:
                - If `data_only` is True, the workbook with calculated values is used.
                Otherwise, the workbook with formulas is used.
                - The constants are extracted from the specified sheet in the workbook,
                starting from the given row and column indices.
                - If `end_row` is not provided, constants will be read until the last row of the sheet.
                - The constants are returned as a DataFrame with constant names as column names and
                constant values in a single row.
            """
        if data_only:
            workbook = self.workbook_values
        else:
            workbook = self.workbook_formula

        if not workbook:
            raise Exception("Workbook is not loaded.")

        sheet = workbook[sheet_name]

        end_col = start_col + 1

        data = list(sheet.iter_rows(min_row=start_row,
                                    min_col=start_col,
                                    max_row=end_row,
                                    max_col=end_col,
                                    values_only=values_only))

        # Create a dictionary to hold the constant names and values
        constants_dict = {row[0]: row[1] for row in data}

        # Create a DataFrame using this dictionary
        data_frame = pd.DataFrame(constants_dict, index=[0])

        return data_frame

    def load_constants_as_data_class(
            self,
            sheet_name: str,
            start_row: int,
            start_col: int,
            end_row: int = None,
            values_only: bool = True,
            data_only: bool = False
    ) -> ConstantsSample:
        """
            Load constants from a specified sheet in the workbook and assign them to a data class instance.

            Parameters:
                - sheet_name (str): The name of the sheet containing the constants.
                - start_row (int): The starting row index for reading constants.
                - start_col (int): The starting column index for reading constants.
                - end_row (int, optional): The ending row index for reading constants.
                 Defaults to None, indicating reading until the last row.
                - values_only (bool, optional): Whether to retrieve only the values of cells. Defaults to True.
                - data_only (bool, optional): Whether to load the workbook with only calculated values.
                Defaults to False.

            Returns:
                ConstantsSample: An instance of the ConstantsSample data class with assigned constants.

            Raises:
                Exception: If the workbook is not loaded.

            Note:
                - If `data_only` is True, the workbook with calculated values is used.
                 Otherwise, the workbook with formulas is used.
                - The constants are extracted from the specified sheet in the workbook,
                starting from the given row and column indices.
                - If `end_row` is not provided, constants will be read until the last row of the sheet.
                - The constants are assigned to the corresponding fields of the ConstantsSample data class instance.
                - The ConstantsSample data class must be defined with fields matching the constant names in the sheet.
            """
        if data_only:
            workbook = self.workbook_values
        else:
            workbook = self.workbook_formula

        if not workbook:
            raise Exception("Workbook is not loaded.")

        sheet = workbook[sheet_name]

        end_col = start_col + 1

        data = list(sheet.iter_rows(min_row=start_row,
                                    min_col=start_col,
                                    max_row=end_row,
                                    max_col=end_col,
                                    values_only=values_only))

        # Create a dictionary to hold the constant names and values
        constants_dict = {row[0]: row[1] for row in data}

        # Instantiate the data class
        data_class_instance = ConstantsSample

        # Assign the constants to the data class fields
        for field in dataclasses.fields(data_class_instance):
            if field.name in constants_dict:
                setattr(data_class_instance, field.name, constants_dict[field.name])

        return data_class_instance

    @staticmethod
    def replace_table_in_specific_sheet_with_data_frame(
            excel_file_path: str,
            sheet_name: str,
            start_row: int,
            header: bool,
            data_frame: pd.DataFrame,
            start_column: int = 1,
    ) -> None:
        """
         Replace a table in a specific sheet of an Excel file with the contents of a DataFrame.

         Parameters:
             - excel_file_path (str): The file path of the Excel file.
             - sheet_name (str): The name of the sheet where the table exists.
             - start_row (int): The starting row index to write the DataFrame contents.
             - header (bool): Whether the DataFrame has a header row.
             - data_frame (pd.DataFrame): The DataFrame containing the data to write.
             - start_column (int, optional): The starting column index to write the DataFrame contents. Defaults to 1.

         Returns:
             None

         Raises:
             Exception: If the excel_file_path is empty or the file does not exist.

         Note:
             - The method opens the Excel file specified by excel_file_path and loads the specified sheet.
             - The existing table in the sheet, starting from the specified start_row and start_column, is removed.
             - The DataFrame contents are written to the worksheet, starting from the specified start_row and start_column.
             - If header=True, the DataFrame's column names are written as the first row.
             - The DataFrame's data is written row by row, with each row corresponding to a row in the worksheet.
             - After writing the DataFrame to the worksheet, the modified workbook is saved back to the original file.

         Example:
             replace_table_in_specific_sheet_with_data_frame(
                 excel_file_path='path/to/excel_file.xlsx',
                 sheet_name='Sheet1',
                 start_row=2,
                 header=True,
                 data_frame=my_data_frame,
                 start_column=1
             )
         """
        if not excel_file_path:
            raise Exception("there is no excel_file with that path")

        workbook = load_workbook(filename=excel_file_path)
        sheet = workbook[sheet_name]

        # change the data frame in to rows. Without index and headers, just only the data.
        rows = dataframe_to_rows(data_frame, index=False, header=header)

        # remove the existing table
        for row in sheet.iter_rows(min_row=start_row, min_col=start_column):
            for cell in row:
                cell.value = None

        # write the new data frame to the worksheet
        for r_idx, row in enumerate(rows):
            for c_idx, value in enumerate(row):
                sheet.cell(row=r_idx + start_row, column=c_idx + start_column, value=value)

        # save copy
        workbook.save(excel_file_path)

    # TODO: does not work yet
    @staticmethod
    def make_copy_of_excel(
            excel_path_to_copy: str,
            output_dir: str,
            output_name: str = "output"
    ):

        # make new path for copy
        # generate a unique filename based on the current date and time
        timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S')
        output_name = f"{output_name}_{timestamp}.xlsx"

        # create the full file path
        output_path = output_dir + '\\' + output_name

        # make copy of excell file
        shutil.copy(excel_path_to_copy, output_path)

    def fill_dict_panda_data_frames(self,
                                    data_frame: pd.DataFrame,
                                    sheets: list[str],
                                    print_process: False | True = False,
                                    ) -> None:
        """
         Fill the dictionary of pandas DataFrames with a single DataFrame for specified sheet names.

         Parameters:
             - data_frame (pd.DataFrame): The DataFrame to be assigned to each sheet name in the dictionary.
             - sheets (list[str]): A list of sheet names to assign the DataFrame.
             - print_process (bool, optional): Whether to print the process for each assigned sheet name.
             Defaults to False.

         Returns:
             None

         Note:
             - The method assigns the provided DataFrame to each sheet name in the dictionary `dict_panda_data_frames`.
             - The DataFrame is assigned to each sheet name as the value in the dictionary.
             - The provided DataFrame is the same for all sheet names in the list.
             - If `print_process` is True, a message will be printed for each assigned sheet name.
         """

        sheet_names = sheets

        for sheet_name in sheet_names:
            self.dict_panda_data_frames[sheet_name] = data_frame

            if print_process is not False:
                print(f'Show process: {sheet_name} is done')

    def get_dict_panda_data_frames(self) -> {str, pd.DataFrame}:
        """
        Get the dictionary of pandas DataFrames.
        Returns:
        dict[str, pd.DataFrame]: The dictionary containing sheet names as keys and corresponding DataFrames as values.
        """
        return self.dict_panda_data_frames

    def fill_dict_constants_data_frames(self,
                                        data_frame: pd.DataFrame,
                                        sheets: list[str],
                                        print_process: False | True = False,
                                        ) -> None:
        """
            Fill the dictionary of constants data frames with a single data frame for specified sheet names.

            Parameters:
                - data_frame (pd.DataFrame): The data frame to be assigned to each sheet name in the dictionary.
                - sheets (list[str]): A list of sheet names to assign the data frame.
                - print_process (bool, optional): Whether to print the process for each assigned sheet name. Defaults to False.

            Returns:
                None

            Note:
                - The method assigns the provided data frame to each sheet name in the dictionary `dict_constants_data_frames`.
                - The data frame is assigned to each sheet name as the value in the dictionary.
                - The provided data frame is the same for all sheet names in the list.
                - If `print_process` is True, a message will be printed for each assigned sheet name.
            """

        sheet_names = sheets

        for sheet_name in sheet_names:
            self.dict_constants_data_frames[sheet_name] = data_frame

            if print_process is not False:
                print(f'Show process: {sheet_name} is done')

    def get_dict_constants_data_frames(self) -> {str, pd.DataFrame}:
        """
        Get the dictionary of constants data frames.
        Returns:
        dict[str, pd.DataFrame]: The dictionary containing sheet names as keys and corresponding data frames as values.
        """
        return self.dict_constants_data_frames

    def fill_dict_constants_data_classes(self,
                                         data_class: ConstantsSample,
                                         sheets: list[str],
                                         print_process: False | True = False,
                                         ) -> None:
        """
            Fill the dictionary of constants data classes with a single data class instance for specified sheet names.

            Parameters:
                - data_class (ConstantsSample): The data class instance to be assigned to each sheet name in the
                dictionary.
                - sheets (list[str]): A list of sheet names to assign the data class instance.
                - print_process (bool, optional): Whether to print the process for each assigned sheet name.
                Defaults to False.

            Returns:
                None

            Note:
                - The method assigns the provided data class instance to each sheet name in the dictionary
                `dict_constants_data_classes`.
                - The data class instance is assigned to each sheet name as the value in the dictionary.
                - The provided data class instance is the same for all sheet names in the list.
                - If `print_process` is True, a message will be printed for each assigned sheet name.
            """
        sheet_names = sheets

        for sheet_name in sheet_names:
            self.dict_constants_data_classes[sheet_name] = data_class

            if print_process is not False:
                print(f'Show process: {sheet_name} is done')

    def get_dict_constants_data_classes(self) -> {str, ConstantsSample}:
        """
        Get the dictionary of constants data classes.
        Returns:
        dict[str, ConstantsSample]: The dictionary containing sheet names as keys and corresponding data class instances as values.
        :return:
        """
        return self.dict_constants_data_classes

    def create_new_workbook(self, data_only: bool = False, **kwargs) -> Workbook:
        """
        Create a new Workbook object by loading an existing workbook file.

        Parameters:
        - data_only (bool, optional): Whether to load the workbook with only calculated values. Defaults to False.
        - **kwargs: Additional keyword arguments to be passed to the `load_workbook` function.

        Returns:
        Workbook: The newly created Workbook object.
        :param data_only:
        :param kwargs:
        :return:
        """
        return load_workbook(filename=self.file_path, data_only=data_only, **kwargs)

    @staticmethod
    def make_excel_based_on_workbook(workbook: Workbook, path_directory: str, filename: str) -> None:
        """
        Save a workbook to the specified directory path.

        Parameters:
            - workbook (Workbook): The workbook to save.
            - path_directory (str): The directory path to save the workbook.
            - filename (str): The file name of the new Excel file.

        Returns:
            None
        """
        filepath = os.path.join(path_directory, filename)
        filepath_with_extension = filepath + ".xlsx"
        workbook.save(filepath_with_extension)
