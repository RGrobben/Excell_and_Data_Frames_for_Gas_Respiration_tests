import unittest

import pandas as pd
from openpyxl import Workbook
from excel_manager import ExcelManager


class TestExcelManager(unittest.TestCase):

    def setUp(self):
        # Create a temporary Excel file for testing
        self.file_path = 'test_file.xlsx'
        self.workbook = Workbook()
        self.workbook.save(self.file_path)

    def tearDown(self):
        # Delete the temporary Excel file after testing
        import os
        os.remove(self.file_path)

    def test_load_workbook(self):
        manager = ExcelManager(self.file_path)
        manager.load_workbook()
        self.assertIsNotNone(manager.workbook_values)
        self.assertEqual(manager.workbook_values.active.title, 'Sheet')

    def test_get_sheet_names(self):
        manager = ExcelManager(self.file_path)
        manager.load_workbook()
        sheet_names = manager.get_sheet_names()
        self.assertEqual(sheet_names, ['Sheet'])

    def test_get_sheet_names_without_loading_workbook(self):
        manager = ExcelManager(self.file_path)
        with self.assertRaises(Exception):
            manager.get_sheet_names()

    def test_load_sheet_table(self):
        # Arrange
        manager = ExcelManager(self.file_path)
        manager.load_workbook()

        # Add test data to the worksheet
        sheet = manager.workbook_values.active
        data = [
            ["header1", "header2"],
            ["data1", "data2"],
            ["data3", "data4"]
        ]
        for i, row in enumerate(data, start=1):
            for j, item in enumerate(row, start=1):
                sheet.cell(row=i, column=j, value=item)

        manager.workbook_values.save(self.file_path)

        # Act
        df = manager.load_sheet_table(sheet_name='Sheet', start_row=1)

        # Assert
        expected_df = pd.DataFrame({"header1": ["data1", "data3"], "header2": ["data2", "data4"]})
        pd.testing.assert_frame_equal(df, expected_df)

    def test_load_sheet_table_without_loading_workbook(self):
        # Arrange
        manager = ExcelManager(self.file_path)

        # Assert
        with self.assertRaises(Exception):
            manager.load_sheet_table(sheet_name='Sheet', start_row=1, start_column=1)

    def test_load_sheet_table_with_invalid_sheet_name(self):
        # Arrange
        manager = ExcelManager(self.file_path)
        manager.load_workbook()

        # Assert
        with self.assertRaises(KeyError):
            manager.load_sheet_table(sheet_name='InvalidSheet', start_row=1, start_column=1)

    def test_load_sheet_table_with_values(self):
        # Arrange
        manager = ExcelManager(self.file_path)
        manager.load_workbook()

        # Add test data to the worksheet
        sheet = manager.workbook_values.active
        data = [
            ["header1", "header2"],
            [10, 20],
            [30, 40],
            [50, 60]
        ]
        for i, row in enumerate(data, start=1):
            for j, item in enumerate(row, start=1):
                sheet.cell(row=i, column=j, value=item)

        manager.workbook_values.save(self.file_path)

        # Act
        df = manager.load_sheet_table(sheet_name='Sheet', start_row=1, start_column=1)

        # Assert
        expected_df = pd.DataFrame({"header1": [10, 30, 50], "header2": [20, 40, 60]})
        pd.testing.assert_frame_equal(df, expected_df)

    def test_load_sheet_table_with_limited_rows(self):
        # Arrange
        manager = ExcelManager(self.file_path)
        manager.load_workbook()

        # Add test data to the worksheet
        sheet = manager.workbook_values.active
        data = [
            ["header1", "header2"],
            ["data1", "data2"],
            ["data3", "data4"],
            ["data5", "data6"],
            ["data7", "data8"],
            ["data9", "data10"],
            ["data11", "data12"],
            ["data13", "data14"],
        ]
        for i, row in enumerate(data, start=1):
            for j, item in enumerate(row, start=1):
                sheet.cell(row=i, column=j, value=item)

        manager.workbook_values.save(self.file_path)

        # Act
        df = manager.load_sheet_table(sheet_name='Sheet', start_row=1, start_column=1, end_row=5)

        # Assert
        expected_df = pd.DataFrame({"header1": ["data1", "data3", "data5", "data7"], "header2": ["data2", "data4", "data6", "data8"]})
        pd.testing.assert_frame_equal(df, expected_df)

    def test_load_sheet_table_with_different_start_column(self):
        # Arrange
        manager = ExcelManager(self.file_path)
        manager.load_workbook()

        # Add test data to the worksheet
        sheet = manager.workbook_values.active
        data = [
            ["header1", "header2", "header3"],
            ["data1", "data2", "data3"],
            ["data4", "data5", "data6"],
            ["data7", "data8", "data9"]
        ]
        for i, row in enumerate(data, start=1):
            for j, item in enumerate(row, start=1):
                sheet.cell(row=i, column=j, value=item)

        manager.workbook_values.save(self.file_path)

        # Act
        df = manager.load_sheet_table(sheet_name='Sheet', start_row=1, start_column=2)

        # Assert
        expected_df = pd.DataFrame({"header2": ["data2", "data5", "data8"], "header3": ["data3", "data6", "data9"]})
        pd.testing.assert_frame_equal(df, expected_df)


if __name__ == '__main__':
    unittest.main()
