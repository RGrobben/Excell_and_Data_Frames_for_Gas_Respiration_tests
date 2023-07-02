from typing import Tuple, Optional

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class NiceExcelFunction:
    """ A class with nice functions for Excel"""
    @staticmethod
    def get_column_index_from_letter(column_letter) -> float | int:
        """
            Get the column index of Excel from a given column letter.

            Parameters:
                - column_letter (str): The column letter to convert to an index.

            Returns:
                int: The corresponding column index.
            Example:
                column_index = WorkbookLoader.get_column_index_from_letter('C')
                # The column_index variable will contain the value 3.
            """
        index = 0
        for char in column_letter:
            index = index * 26 + (ord(char.upper()) - ord('A')) + 1
        return index

    @staticmethod
    def get_column_letter_from_index(index) -> str:
        """
            Get the column letter of Excel from a given column index.

            Parameters:
                - index (int): The column index to convert to a letter.

            Returns:
                str: The corresponding column letter.

            Example:
                column_letter = WorkbookLoader.get_column_letter_from_index(3)
                # The column_letter variable will contain the value 'C'.
            """
        letters = ''
        while index:
            index, remainder = divmod(index - 1, 26)
            letters = chr(65 + remainder) + letters
        return letters

    @staticmethod
    def find_string_index(excel_file_path: str,
                          sheet_name: str,
                          search_string: str,
                          max_iterations: int = 1000
                          ) -> Tuple[Optional[int], Optional[int]]:
        """
        Finds the index (row and column) in a sheet of an Excel file or workbook for a specific string.

        Args:
            excel_file_path (str): Path to the Excel file.
            sheet_name (str): Name of the sheet to search in.
            search_string (str): String to search for.
            max_iterations (int, optional): Maximum number of iterations to perform. Defaults to 1000.

        Returns:
            Tuple[int, int]: Row and column indices of the found cell, or (None, None) if the search string is not found.
        """
        workbook = load_workbook(excel_file_path)
        sheet = workbook[sheet_name]

        iteration_count = 0

        # Iterate over all cells in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == search_string:
                    # Return the row and column indices of the found cell
                    return cell.row, cell.column

                iteration_count += 1

                if iteration_count >= max_iterations:
                    # Reached the maximum number of iterations
                    return None, None

    @staticmethod
    def find_column_name_excel_index_based_on_column_name_string_in_given_row(workbook,
                                                                              sheet_name: str,
                                                                              search_string: str,
                                                                              header_row: int,
                                                                              ) -> str:
        """
            Find the Excel column name (letter) based on the column name string in the given header row.

            Parameters:
                - workbook (Workbook): The workbook containing the sheet.
                - sheet_name (str): The name of the sheet in which to search.
                - search_string (str): The column name string to search for in the header row.
                - header_row (int): The row index of the header row.

            Returns:
                str: The Excel column name (letter) corresponding to the column name string.

            Raises:
                ValueError: If the search string is not found in the header row.
            """

        # Load the workbook and select the sheet
        sheet = workbook[sheet_name]

        # Get the maximum column index
        max_column = sheet.max_column

        # Iterate over the cells in the header row
        for column in range(1, max_column + 1):
            cell = sheet.cell(row=header_row, column=column)
            cell_value = str(cell.value)

            # Check if the search string is found in the cell value
            if search_string == cell_value:
                return get_column_letter(cell.column)

        # If the search string is not found, raise an exception
        raise ValueError(f"Search string '{search_string}' not found in the header row.")

    @staticmethod
    def get_cell_value_pandas(file_path, sheet_name, cell) -> float | int:
        """
            Get the value of a specific cell in an Excel file using pandas.

            Parameters:
                - file_path (str): The path to the Excel file.
                - sheet_name (str): The name of the sheet containing the cell.
                - cell (str): The cell reference (e.g., 'A1', 'B2').

            Returns:
                object: The value of the specified cell.

            Example:
                cell_value = WorkbookLoader.get_cell_value_pandas(file_path='path/to/excel_file.xlsx',
                                                                 sheet_name='Sheet1',
                                                                 cell='A1')
                # The cell_value variable will contain the value of cell A1 in Sheet1.
            """
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        value = df.loc[cell].values[0]
        return value

    @staticmethod
    def get_cell_value_openpyxl(file_path, sheet_name, cell) -> float | int:
        """
            Get the value of a specific cell in an Excel file using openpyxl.

            Parameters:
                - file_path (str): The path to the Excel file.
                - sheet_name (str): The name of the sheet containing the cell.
                - cell (str): The cell reference (e.g., 'A1', 'B2').

            Returns:
                object: The value of the specified cell.

            Example:
                cell_value = WorkbookLoader.get_cell_value_openpyxl(file_path='path/to/excel_file.xlsx',
                                                                    sheet_name='Sheet1',
                                                                    cell='A1')
                # The cell_value variable will contain the value of cell A1 in Sheet1.
            """
        wb = load_workbook(filename=file_path)
        sheet = wb[sheet_name]
        value = sheet[cell].value
        return value


class NicePandaFrameFunctions:
    """ A class with nice functions for panda data frames"""
    @staticmethod
    def get_cell_value_pandas_df(dataframe: pd.DataFrame, row: int, column: int) -> str | float:
        """
            Get the value of a specific cell in a pandas DataFrame.

            Parameters:
                - dataframe (pd.DataFrame): The DataFrame containing the cell.
                - row (int): The row index of the cell.
                - column (int): The column index of the cell.

            Returns:
                str or float: The value of the specified cell.

            Example:
                my_dataframe = pd.DataFrame(...)
                cell_value = NicePandaFrameFunctions.get_cell_value_pandas_df(dataframe=my_dataframe, row=0, column=0)
                # The cell_value variable will contain the value of the cell at row 0, column 0 in the DataFrame.
            """
        value = dataframe.loc[row:column]
        return value
