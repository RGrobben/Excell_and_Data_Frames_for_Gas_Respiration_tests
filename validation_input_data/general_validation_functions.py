from datetime import datetime
from typing import Dict, Any, List

import pandas as pd
from openpyxl.styles import PatternFill

from nice_functions import NiceExcelFunction


def validate_if_there_is_a_float_or_integer_in_cell(data_frame: pd.DataFrame, column_name: str,
                                                    start_row_values_table_in_excel: int = 0) -> tuple:
    invalid_rows = []
    column_in_data_frame_to_be_checked = data_frame[column_name]
    for index in column_in_data_frame_to_be_checked.index:
        value = column_in_data_frame_to_be_checked[index]
        if not isinstance(value, (float, int)) or pd.isnull(value):
            invalid_rows.append(index + start_row_values_table_in_excel)

    if len(invalid_rows) > 0:
        return column_name, invalid_rows
    else:
        return column_name, True


def validate_if_there_is_no_specific_float_or_integer_in_cell(data_frame: pd.DataFrame, column_name: str,
                                                              specific_float_or_integer: float | int,
                                                              start_row_values_table_in_excel: int = 0) -> tuple:
    invalid_rows = []
    column_in_data_frame_to_be_checked = data_frame[column_name]
    for index in column_in_data_frame_to_be_checked.index:
        value = column_in_data_frame_to_be_checked[index]
        if not value == specific_float_or_integer or pd.isnull(value):
            invalid_rows.append(index + start_row_values_table_in_excel)

    if len(invalid_rows) > 0:
        return column_name, invalid_rows
    else:
        return column_name, True


def validate_if_there_is_a_string(data_frame: pd.DataFrame, column_name: str,
                                  start_row_values_table_in_excel: int = 0) -> tuple:
    invalid_rows = []
    column_in_data_frame_to_be_checked = data_frame[column_name]
    for index in column_in_data_frame_to_be_checked.index:
        value = column_in_data_frame_to_be_checked[index]
        if not isinstance(value, int) or pd.isnull(value):
            invalid_rows.append(index + start_row_values_table_in_excel)

    if len(invalid_rows) > 0:
        return column_name, invalid_rows
    else:
        return column_name, True


def validate_if_there_is_a_specific_string(data_frame: pd.DataFrame, column_name: str, specific_string: str,
                                           start_row_values_table_in_excel: int = 0) -> tuple:
    invalid_rows = []
    column_in_data_frame_to_be_checked = data_frame[column_name]
    for index in column_in_data_frame_to_be_checked.index:
        value = column_in_data_frame_to_be_checked[index]
        if pd.isnull(value) or value == specific_string:
            invalid_rows.append(index + start_row_values_table_in_excel)

    if len(invalid_rows) > 0:
        return column_name, invalid_rows
    else:
        return column_name, True


def validate_if_there_is_in_cell_one_of_the_specific_strings(data_frame: pd.DataFrame, column_name: str,
                                                             list_specific_string: [str],
                                                             start_row_values_table_in_excel: int = 0) -> tuple:
    invalid_rows = []
    column_in_data_frame_to_be_checked = data_frame[column_name]
    for index in column_in_data_frame_to_be_checked.index:
        value = column_in_data_frame_to_be_checked[index]
        if pd.isnull(value) or value not in list_specific_string:
            invalid_rows.append(index + start_row_values_table_in_excel)

    if len(invalid_rows) > 0:
        return column_name, invalid_rows
    else:
        return column_name, True


def check_is_string_date(string: str, date_format) -> bool:
    """
    Check if a string can be parsed as a valid date using the given format.

    Parameters:
        string (str): The string to be checked for date validity.
        date_format (str): The format of the date values. For example: "%Y-%m-%d"

    Returns:
        bool: True if the string is a valid date, False otherwise.
    """
    try:
        datetime.strptime(string, date_format)
        return True
    except ValueError:
        return False


def validate_if_in_cell_is_correct_date_or_not_filled(data_frame: pd.DataFrame, column_name_date: str,
                                                      format_date: str = "%Y-%m-%d",
                                                      start_row_values_table_in_excel: int = 0) -> tuple:
    """
        Validate the indexes of a pandas DataFrame for incorrect date values in a specific column.

        Parameters:
            data_frame (pd.DataFrame): The pandas DataFrame to be validated.
            column_name_date (str): The name of the column containing date values to be checked.
            format_date (str, optional): The format of the date values. Defaults to "%Y-%m-%d".
            start_row_values_table_in_excel (int, optional): The starting row index of the table in Excel. Defaults to 0.

        Returns:
            tuple: A tuple containing the column name and either the list of invalid row indexes or True if all rows are valid.
        """

    invalid_rows = []
    column_in_data_frame_to_be_checked = data_frame[column_name_date].astype(str)
    for index in column_in_data_frame_to_be_checked.index:
        value = column_in_data_frame_to_be_checked[index]

        if pd.isnull(value) or check_is_string_date(string=value, date_format=format_date) is False:
            invalid_rows.append(index + start_row_values_table_in_excel)

    if len(invalid_rows) > 0:
        return column_name_date, invalid_rows
    else:
        return column_name_date, True


def check_is_string_time(string: str, time_format) -> bool:
    """
    Check if a string can be parsed as a valid time using the given format.

    Parameters:
        string (str): The string to be checked for time validity.
        time_format (str): The format of the time values. For example: "%H:%M:%S"

    Returns:
        bool: True if the string is a valid time, False otherwise.
    """
    try:
        datetime.strptime(string, time_format)
        return True
    except ValueError:
        return False


def validate_if_in_cell_is_correct_time_or_not_filled(data_frame: pd.DataFrame, column_name_time: str,
                                                      format_time: str = "%H:%M:%S",
                                                      start_row_values_table_in_excel: int = 0) -> tuple:
    """
    Valideer de indexen van een pandas DataFrame op onjuiste tijdwaarden in een specifieke kolom.

    Parameters:
        data_frame (pd.DataFrame): De pandas DataFrame die moet worden gevalideerd.
        column_name_time (str): De naam van de kolom met tijdwaarden die gecontroleerd moeten worden.
        format_time (str, optioneel): Het formaat van de tijdwaarden. Standaard is "%H:%M:%S".
        start_row_values_table_in_excel (int, optioneel): De startindex van de tabel in Excel. Standaard is 0.

    Returns:
        tuple: Een tuple met de kolomnaam en ofwel de lijst met ongeldige rij-indexen of True als alle rijen geldig zijn.
    """

    invalid_rows = []
    column_in_data_frame_to_be_checked = data_frame[column_name_time].astype(str)
    for index in column_in_data_frame_to_be_checked.index:
        value = column_in_data_frame_to_be_checked[index]

        if pd.isnull(value) or check_is_string_time(string=value, time_format=format_time) is False:
            invalid_rows.append(index + start_row_values_table_in_excel)

    if len(invalid_rows) > 0:
        return column_name_time, invalid_rows
    else:
        return column_name_time, True


def style_color_cells_with_given_indexes(workbook, dict_sheet_name_column_names_indexes: {},
                                         header_row: int,
                                         color: str, fill_type: str,
                                         start_row_values_table_in_excel: int = 1,
                                         show_process: bool = False) -> None:
    """

    :param show_process: when you like to see the process.
    :param workbook: The workbook
    :param dict_sheet_name_column_names_indexes: {sheet_name: {column_name: indexes}}
    :param header_row: the row where the header is.
    :param color: the colorcode based on RGB colors
    :param fill_type: Fill type: "solid", "gradient", "patter", "None"
    :param start_row_values_table_in_excel: the starting row of the values of the table in Excel. In Excel the rows
    are starting with 1.
    """
    color_fill = PatternFill(start_color=color, end_color=color, fill_type=fill_type)

    for sheet_name, sheet_data in dict_sheet_name_column_names_indexes.items():
        sheet = workbook[sheet_name]
        for column_name, indexes in sheet_data.items():
            column_letter = NiceExcelFunction.find_column_name_excel_index_based_on_column_name_string_in_given_row(
                workbook=workbook, sheet_name=sheet_name, search_string=column_name, header_row=header_row)
            for index in indexes:
                index_row_in_excel = start_row_values_table_in_excel + index
                column_and_row_excel_combination = column_letter + str(index_row_in_excel)
                sheet[column_and_row_excel_combination].fill = color_fill

        if show_process is True:
            print(f"Finished with coloring {sheet_name}")


def style_color_cells_with_given_excel_indexes_and_excel_column_name(workbook, dict_sheet_name_indexes: {},
                                                                     excel_column_name: str,
                                                                     color: str,
                                                                     fill_type: str,
                                                                     show_process: bool = False) -> None:
    color_fill = PatternFill(start_color=color, end_color=color, fill_type=fill_type)

    for sheet_name, indexes in dict_sheet_name_indexes.items():
        sheet = workbook[sheet_name]
        if show_process is True:
            print(f"started with {sheet_name}")
        for index in indexes:
            index_row_in_excel = index
            column_and_row_excel_combination = excel_column_name + str(index_row_in_excel)
            sheet[column_and_row_excel_combination].fill = color_fill
